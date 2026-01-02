"""
Portfolio report renderer.

This module provides functions to generate portfolio reports in multiple formats
(PDF, CSV, Excel) from valuation runs and exposure data.

Key functions:
- generate_portfolio_report: Main entry point for report generation
- render_pdf_report: Generate PDF report from HTML template
- render_csv_report: Generate CSV report with exposure tables
- render_excel_report: Generate Excel report with multiple sheets
"""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from django.core.files.base import ContentFile
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from weasyprint import HTML

from apps.analytics.models import ExposureDimensionType, RunStatus, ValuationRun
from apps.reports.models import Report, ReportStatus, ReportTemplate


def generate_portfolio_report(valuation_run_id: int) -> Report:
    """
    Generate portfolio report for a valuation run.

    Main entry point that creates a Report record and generates PDF, CSV, and Excel
    files. Updates the report status throughout the process.

    Args:
        valuation_run_id: ValuationRun ID to generate report for.

    Returns:
        Report instance with generated files.

    Raises:
        ValuationRun.DoesNotExist: If valuation run doesn't exist.
        ReportTemplate.DoesNotExist: If default template doesn't exist.

    Example:
        >>> run = ValuationRun.objects.get(id=1)
        >>> report = generate_portfolio_report(run.id)
        >>> print(report.pdf_file.url)
    """
    # Get valuation run
    valuation_run = ValuationRun.objects.select_related(
        "portfolio", "organization"
    ).get(id=valuation_run_id)

    # Validate that run is successful
    if valuation_run.status != RunStatus.SUCCESS:
        raise ValueError(
            f"Cannot generate report for run with status {valuation_run.status}. "
            "Run must be in SUCCESS status."
        )

    # Get or create default template (for MVP, use a simple default)
    template, _ = ReportTemplate.objects.get_or_create(
        organization=valuation_run.organization,
        name="Portfolio Overview v1",
        version="1.0",
        defaults={
            "template_type": "portfolio_overview",
            "config_json": {
                "sections": ["overview", "exposures", "concentration", "data_quality"]
            },
            "is_active": True,
        },
    )

    # Create report record
    report = Report.objects.create(
        valuation_run=valuation_run,
        template=template,
        status=ReportStatus.GENERATING,
        organization=valuation_run.organization,
    )

    try:
        # Generate PDF
        pdf_bytes = render_pdf_report(valuation_run, template)
        report.pdf_file.save(
            f"portfolio_report_{valuation_run.as_of_date}.pdf",
            ContentFile(pdf_bytes),
            save=False,
        )

        # Generate CSV
        csv_bytes = render_csv_report(valuation_run)
        report.csv_file.save(
            f"portfolio_report_{valuation_run.as_of_date}.csv",
            ContentFile(csv_bytes),
            save=False,
        )

        # Generate Excel
        excel_bytes = render_excel_report(valuation_run)
        report.excel_file.save(
            f"portfolio_report_{valuation_run.as_of_date}.xlsx",
            ContentFile(excel_bytes),
            save=False,
        )

        # Update status and timestamp
        report.status = ReportStatus.SUCCESS
        report.generated_at = timezone.now()
        report.save()

    except Exception as e:
        # Update status with error
        report.status = ReportStatus.FAILED
        report.error_message = str(e)
        report.save()
        raise

    return report


def render_pdf_report(valuation_run: ValuationRun, template: ReportTemplate) -> bytes:
    """
    Render PDF report from HTML template.

    Renders the HTML template using Django's template system and converts it to PDF
    using WeasyPrint. Returns PDF bytes.

    Args:
        valuation_run: ValuationRun instance.
        template: ReportTemplate instance.

    Returns:
        PDF file content as bytes.

    Example:
        >>> run = ValuationRun.objects.get(id=1)
        >>> template = ReportTemplate.objects.get(name="Portfolio Overview v1")
        >>> pdf_bytes = render_pdf_report(run, template)
        >>> with open('report.pdf', 'wb') as f:
        ...     f.write(pdf_bytes)
    """
    # Prepare template context
    context = _prepare_template_context(valuation_run)

    # Render HTML template
    html_string = render_to_string("reports/portfolio_overview_v1.html", context)

    # Convert HTML to PDF using WeasyPrint
    html_doc = HTML(string=html_string, base_url=None)
    pdf_bytes = html_doc.write_pdf()

    return pdf_bytes


def render_csv_report(valuation_run: ValuationRun) -> bytes:
    """
    Render CSV report with exposure tables.

    Generates a CSV file with exposure breakdowns across all dimensions.
    Uses standard library csv module.

    Args:
        valuation_run: ValuationRun instance.

    Returns:
        CSV file content as bytes.

    Example:
        >>> run = ValuationRun.objects.get(id=1)
        >>> csv_bytes = render_csv_report(run)
        >>> with open('report.csv', 'wb') as f:
        ...     f.write(csv_bytes)
    """
    # Use StringIO for CSV writer (text-based), then encode to bytes
    # Use StringIO for CSV writer (text-based), then encode to bytes
    output = StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(["Portfolio Report", valuation_run.portfolio.name])
    writer.writerow(["As-of Date", str(valuation_run.as_of_date)])
    writer.writerow(["Total Market Value", str(valuation_run.get_total_market_value())])
    writer.writerow([])

    # Currency exposures
    writer.writerow(["Currency Exposures"])
    writer.writerow(["Currency", "Value (Base Currency)", "Percentage"])
    currency_exposures = valuation_run.get_exposures(ExposureDimensionType.CURRENCY)
    for exp in currency_exposures:
        writer.writerow(
            [exp.dimension_label, str(exp.value_base), f"{exp.pct_total:.2f}%"]
        )
    writer.writerow([])

    # Issuer exposures (top 10)
    writer.writerow(["Top Issuer Exposures"])
    writer.writerow(["Issuer", "Value (Base Currency)", "Percentage"])
    issuer_exposures = valuation_run.get_exposures(
        ExposureDimensionType.ISSUER
    ).order_by("-value_base")[:10]
    for exp in issuer_exposures:
        writer.writerow(
            [exp.dimension_label, str(exp.value_base), f"{exp.pct_total:.2f}%"]
        )
    writer.writerow([])

    # Country exposures
    writer.writerow(["Country Exposures"])
    writer.writerow(["Country", "Value (Base Currency)", "Percentage"])
    country_exposures = valuation_run.get_exposures(ExposureDimensionType.COUNTRY)
    for exp in country_exposures:
        writer.writerow(
            [exp.dimension_label, str(exp.value_base), f"{exp.pct_total:.2f}%"]
        )

    # Convert string output to bytes (UTF-8 encoding)
    output.seek(0)
    return output.getvalue().encode("utf-8")


def render_excel_report(valuation_run: ValuationRun) -> bytes:
    """
    Render Excel report with multiple sheets.

    Generates an Excel file with multiple sheets:
    - Overview: Portfolio summary
    - Exposures: Currency, issuer, country exposures
    - Concentration: Top concentrations
    - Data Quality: Data quality summary

    Args:
        valuation_run: ValuationRun instance.

    Returns:
        Excel file content as bytes.

    Example:
        >>> run = ValuationRun.objects.get(id=1)
        >>> excel_bytes = render_excel_report(run)
        >>> with open('report.xlsx', 'wb') as f:
        ...     f.write(excel_bytes)
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Overview sheet
    overview_sheet = wb.create_sheet("Overview")
    overview_sheet.append(["Portfolio Report"])
    overview_sheet.append(["Portfolio", valuation_run.portfolio.name])
    overview_sheet.append(["As-of Date", str(valuation_run.as_of_date)])
    overview_sheet.append(
        ["Total Market Value", str(valuation_run.get_total_market_value())]
    )
    overview_sheet.append(["Position Count", valuation_run.position_count])
    overview_sheet.append([])

    # Currency exposures sheet
    currency_sheet = wb.create_sheet("Currency Exposures")
    currency_sheet.append(["Currency", "Value (Base Currency)", "Percentage"])
    currency_exposures = valuation_run.get_exposures(ExposureDimensionType.CURRENCY)
    for exp in currency_exposures:
        currency_sheet.append(
            [exp.dimension_label, float(exp.value_base.amount), float(exp.pct_total)]
        )

    # Issuer exposures sheet (top 20)
    issuer_sheet = wb.create_sheet("Issuer Exposures")
    issuer_sheet.append(["Issuer", "Value (Base Currency)", "Percentage"])
    issuer_exposures = valuation_run.get_exposures(
        ExposureDimensionType.ISSUER
    ).order_by("-value_base")[:20]
    for exp in issuer_exposures:
        issuer_sheet.append(
            [exp.dimension_label, float(exp.value_base.amount), float(exp.pct_total)]
        )

    # Country exposures sheet
    country_sheet = wb.create_sheet("Country Exposures")
    country_sheet.append(["Country", "Value (Base Currency)", "Percentage"])
    country_exposures = valuation_run.get_exposures(ExposureDimensionType.COUNTRY)
    for exp in country_exposures:
        country_sheet.append(
            [exp.dimension_label, float(exp.value_base.amount), float(exp.pct_total)]
        )

    # Data quality sheet
    dq_sheet = wb.create_sheet("Data Quality")
    dq_summary = valuation_run.get_data_quality_summary()
    dq_sheet.append(["Metric", "Value"])
    dq_sheet.append(["Total Positions", dq_summary["total_positions"]])
    dq_sheet.append(["Positions with Issues", dq_summary["positions_with_issues"]])
    dq_sheet.append(["Missing FX Rates", dq_summary["missing_fx_rates"]])
    dq_sheet.append(["Invalid FX Rates", dq_summary["invalid_fx_rates"]])

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _prepare_template_context(valuation_run: ValuationRun) -> dict:
    """
    Prepare template context for report rendering.

    Args:
        valuation_run: ValuationRun instance.

    Returns:
        Dictionary with template context variables.
    """
    from apps.analytics.engine.exposures import compute_top_concentrations

    # Get exposures by dimension type
    exposures = {
        "currency": list(
            valuation_run.get_exposures(ExposureDimensionType.CURRENCY).order_by(
                "-value_base"
            )[:10]
        ),
        "issuer": list(
            valuation_run.get_exposures(ExposureDimensionType.ISSUER).order_by(
                "-value_base"
            )[:10]
        ),
        "country": list(
            valuation_run.get_exposures(ExposureDimensionType.COUNTRY).order_by(
                "-value_base"
            )[:10]
        ),
        "instrument_group": list(
            valuation_run.get_exposures(
                ExposureDimensionType.INSTRUMENT_GROUP
            ).order_by("-value_base")
        ),
        "instrument_type": list(
            valuation_run.get_exposures(ExposureDimensionType.INSTRUMENT_TYPE).order_by(
                "-value_base"
            )
        ),
    }

    # Get top concentrations
    results = valuation_run.get_results().select_related(
        "position_snapshot__instrument__issuer",
        "position_snapshot__instrument__instrument_group",
        "position_snapshot__instrument__instrument_type",
    )
    total_mv = valuation_run.get_total_market_value()

    concentrations = {
        "issuers": compute_top_concentrations(results, "issuer", total_mv, top_n=5),
        "instruments": compute_top_concentrations(
            results, "instrument", total_mv, top_n=5
        ),
    }

    # Get data quality summary
    data_quality = valuation_run.get_data_quality_summary()

    return {
        "portfolio": valuation_run.portfolio,
        "valuation_run": valuation_run,
        "exposures": exposures,
        "concentrations": concentrations,
        "data_quality": data_quality,
        "as_of_date": valuation_run.as_of_date,
    }
