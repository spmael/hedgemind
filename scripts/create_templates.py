"""
Script to create Excel templates for portfolio imports and reference data.

This script generates standardized Excel templates that serve as guidance
for users importing data into Hedgemind. Templates are not enforced but
provide recommended column names and example data.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Template directory
TEMPLATE_DIR = Path(__file__).parent.parent / "docs" / "templates"
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)


def create_portfolio_holdings_template():
    """Create portfolio holdings template with canonical column names."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    # Headers
    headers = [
        "instrument_identifier",
        "quantity",
        "currency",
        "price",
        "market_value",
        "book_value",
        "valuation_source",
        "accrued_interest",
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Example data rows
    ws.append(["CM1234567890", 1000, "XAF", 105.50, 1055000, 1000000, "custodian", 0])
    ws.append(["CM9876543210", 500, "XAF", 98.25, 491250, 500000, "market", 1250])
    ws.append(["EQUITY001", 2500, "XAF", 45.00, 112500, 100000, "internal", 0])

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.append(["Field", "Description", "Required", "Example"])
    ws_instructions.append(
        [
            "instrument_identifier",
            "ISIN or ticker symbol",
            "Yes",
            "CM1234567890 or EQUITY001",
        ]
    )
    ws_instructions.append(["quantity", "Number of units/shares", "Yes", "1000"])
    ws_instructions.append(["currency", "ISO currency code (3 letters)", "Yes", "XAF"])
    ws_instructions.append(
        ["price", "Price per unit (optional if market_value provided)", "No*", "105.50"]
    )
    ws_instructions.append(
        [
            "market_value",
            "Total market value (optional if price provided)",
            "No*",
            "1055000",
        ]
    )
    ws_instructions.append(["book_value", "Book/cost value", "Yes", "1000000"])
    ws_instructions.append(
        [
            "valuation_source",
            "Source of valuation (custodian, market, internal, manual)",
            "Yes",
            "custodian",
        ]
    )
    ws_instructions.append(
        ["accrued_interest", "Accrued interest (for bonds)", "No", "1250"]
    )
    ws_instructions.append(
        [
            "",
            "* At least one of 'price' or 'market_value' must be provided",
            "",
            "",
        ]
    )

    # Style instructions header
    for cell in ws_instructions[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Auto-adjust column widths for instructions
    for col in range(1, 5):
        ws_instructions.column_dimensions[get_column_letter(col)].width = 30

    wb.save(TEMPLATE_DIR / "portfolio_holdings_template.xlsx")
    print(f"Created {TEMPLATE_DIR / 'portfolio_holdings_template.xlsx'}")


def create_instrument_template():
    """Create instrument master template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "INSTRUMENTS"

    # Headers (required first, then optional)
    headers = [
        "name",
        "instrument_group_code",
        "instrument_type_code",
        "currency",
        "issuer_code",
        "valuation_method",
        "isin",
        "ticker",
        "country",
        "sector",
        "maturity_date",
        "coupon_rate",
        "coupon_frequency",
        "first_listing_date",
        "original_offering_amount",
        "units_outstanding",
        "face_value",
        "amortization_method",
        "last_coupon_date",
        "next_coupon_date",
        "fund_category",
        "fund_launch_date",
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Example data rows
    ws.append(
        [
            "Cameroon 5Y Government Bond",
            "Government Bonds",
            "Bond",
            "XAF",
            "REP_CAMEROON",
            "mark_to_market",
            "CM1234567890",
            "CMBOND5Y",
            "CM",
            "Government",
            "2029-12-31",
            5.5,
            "ANNUAL",
            "2024-01-15",
            100000000.0,
            None,
            1000.0,
            "BULLET",
            "2024-12-31",
            "2025-12-31",
            None,
            None,
        ]
    )
    ws.append(
        [
            "Equity Stock Example",
            "Equities",
            "Equity",
            "XAF",
            "CORP_EXAMPLE",
            "mark_to_market",
            None,
            "EQUITY001",
            "CM",
            "Financial Services",
            None,
            None,
            None,
            "2020-01-01",
            None,
            1000000.0,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.append(["Field", "Description", "Required", "Example"])
    ws_instructions.append(
        ["name", "Full name of instrument", "Yes", "Cameroon 5Y Government Bond"]
    )
    ws_instructions.append(
        [
            "instrument_group_code",
            "Code of InstrumentGroup (must exist)",
            "Yes",
            "Government Bonds",
        ]
    )
    ws_instructions.append(
        [
            "instrument_type_code",
            "Code of InstrumentType within group (must exist)",
            "Yes",
            "Bond",
        ]
    )
    ws_instructions.append(["currency", "ISO currency code (3 letters)", "Yes", "XAF"])
    ws_instructions.append(
        [
            "issuer_code",
            "Issuer short_name or name (must exist)",
            "Yes",
            "REP_CAMEROON",
        ]
    )
    ws_instructions.append(
        [
            "valuation_method",
            "mark_to_market, mark_to_model, external_appraisal, manual_declared",
            "Yes",
            "mark_to_market",
        ]
    )
    ws_instructions.append(["isin", "ISIN code (optional)", "No", "CM1234567890"])
    ws_instructions.append(["ticker", "Ticker symbol (optional)", "No", "CMBOND5Y"])
    ws_instructions.append(["country", "2-letter country code (optional)", "No", "CM"])
    ws_instructions.append(["sector", "Economic sector (optional)", "No", "Government"])
    ws_instructions.append(
        ["maturity_date", "Maturity date YYYY-MM-DD (optional)", "No", "2029-12-31"]
    )
    ws_instructions.append(
        ["coupon_rate", "Coupon rate as percentage (optional)", "No", "5.5"]
    )
    ws_instructions.append(
        ["coupon_frequency", "ANNUAL, SEMI_ANNUAL, etc. (optional)", "No", "ANNUAL"]
    )
    ws_instructions.append(
        [
            "first_listing_date",
            "First listing date YYYY-MM-DD (optional)",
            "No",
            "2024-01-15",
        ]
    )
    ws_instructions.append(
        [
            "original_offering_amount",
            "Original offering amount (optional)",
            "No",
            "100000000.0",
        ]
    )
    ws_instructions.append(
        ["units_outstanding", "Units/shares outstanding (optional)", "No", "1000000.0"]
    )
    ws_instructions.append(["face_value", "Face/par value (optional)", "No", "1000.0"])
    ws_instructions.append(
        [
            "amortization_method",
            "BULLET, AMORTIZING, ZERO_COUPON (optional)",
            "No",
            "BULLET",
        ]
    )
    ws_instructions.append(
        [
            "last_coupon_date",
            "Last coupon date YYYY-MM-DD (optional)",
            "No",
            "2024-12-31",
        ]
    )
    ws_instructions.append(
        [
            "next_coupon_date",
            "Next coupon date YYYY-MM-DD (optional)",
            "No",
            "2025-12-31",
        ]
    )
    ws_instructions.append(
        [
            "fund_category",
            "DIVERSIFIED, MONEY_MARKET, BOND, EQUITY (optional, for funds)",
            "No",
            "DIVERSIFIED",
        ]
    )
    ws_instructions.append(
        [
            "fund_launch_date",
            "Fund launch date YYYY-MM-DD (optional)",
            "No",
            "2020-01-01",
        ]
    )

    # Style instructions header
    for cell in ws_instructions[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Auto-adjust column widths for instructions
    for col in range(1, 5):
        ws_instructions.column_dimensions[get_column_letter(col)].width = 30

    wb.save(TEMPLATE_DIR / "instrument_master_template.xlsx")
    print(f"Created {TEMPLATE_DIR / 'instrument_master_template.xlsx'}")


def create_issuer_template():
    """Create issuer master template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ISSUERS"

    # Headers
    headers = ["name", "short_name", "country", "issuer_group"]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Example data rows
    ws.append(["Republic of Cameroon", "REP_CAMEROON", "CM", "Sovereign"])
    ws.append(["Africa Bright Asset Management", "ABAM", "GA", "Asset Manager"])
    ws.append(["Example Corporation", "CORP_EXAMPLE", "CM", "Corporate"])

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.append(["Field", "Description", "Required", "Example"])
    ws_instructions.append(
        [
            "name",
            "Full legal name of issuer (unique per organization)",
            "Yes",
            "Republic of Cameroon",
        ]
    )
    ws_instructions.append(
        [
            "short_name",
            "Short name or abbreviation (used as issuer_code)",
            "Yes",
            "REP_CAMEROON",
        ]
    )
    ws_instructions.append(["country", "2-letter ISO country code", "Yes", "CM"])
    ws_instructions.append(
        [
            "issuer_group",
            "Group classification (e.g., Sovereign, Corporate)",
            "Yes",
            "Sovereign",
        ]
    )

    # Style instructions header
    for cell in ws_instructions[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Auto-adjust column widths for instructions
    for col in range(1, 5):
        ws_instructions.column_dimensions[get_column_letter(col)].width = 30

    wb.save(TEMPLATE_DIR / "issuer_master_template.xlsx")
    print(f"Created {TEMPLATE_DIR / 'issuer_master_template.xlsx'}")


if __name__ == "__main__":
    print("Creating Excel templates...")
    create_portfolio_holdings_template()
    create_instrument_template()
    create_issuer_template()
    print("All templates created successfully!")
